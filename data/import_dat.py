import sqlite3
import openpyxl
import re

conn = sqlite3.connect('peptides.db')
cursor = conn.cursor()

wb = openpyxl.load_workbook('1.xlsx')
sheet = wb.active

# 打印原始列名查看
cols = [cell.value for cell in sheet[1]]
print("Original columns:", cols)

# 清理并替换 None 值
cols = [re.sub(r'\W+', '_', str(cell.value or f'column_{i}')) for i, cell in enumerate(sheet[1])]
print("Cleaned columns:", cols)

create_table = f'''
CREATE TABLE peptides (
    {','.join([f'{col} TEXT' for col in cols])}
)'''

print("SQL:", create_table)

import sqlite3
import openpyxl
import re

conn = sqlite3.connect('peptides.db')
cursor = conn.cursor()

wb = openpyxl.load_workbook('1.xlsx')
sheet = wb.active

# 获取并清理列名
cols = [re.sub(r'\W+', '_', str(cell.value or f'column_{i}')) for i, cell in enumerate(sheet[1])]

create_table = f'''
CREATE TABLE peptides (
{','.join([f'{col} TEXT' for col in cols])}
)'''

cursor.execute(create_table)

# 插入数据
for row in sheet.iter_rows(min_row=2):
    values = [str(cell.value or '') for cell in row]
    q_marks = ','.join(['?' for _ in values])
    cursor.execute(f'INSERT INTO peptides VALUES ({q_marks})', values)

conn.commit()
conn.close()

import sqlite3
import openpyxl

# 连接数据库
conn = sqlite3.connect('peptides.db')
cursor = conn.cursor()

# 删除已存在的表
cursor.execute('DROP TABLE IF EXISTS peptides')

# 读取Excel
wb = openpyxl.load_workbook('1.xlsx')
sheet = wb.active

# 获取非空列名
valid_cols = [col.value for col in sheet[1] if col.value is not None]
cols = [col.replace(' ', '_').replace('/', '_').replace('(','').replace(')','') for col in valid_cols]

print("Creating table with columns:", cols)

# 创建表
create_table = f'''CREATE TABLE peptides (
   {','.join([f'"{col}" TEXT' for col in cols])}
)'''
cursor.execute(create_table)

# 插入数据
for row in sheet.iter_rows(min_row=2):
   values = [str(cell.value or '') for cell in row[:len(valid_cols)]]
   placeholders = ','.join(['?' for _ in values])
   cursor.execute(f'INSERT INTO peptides VALUES ({placeholders})', values)

conn.commit()
conn.close()
print("Data imported successfully!")